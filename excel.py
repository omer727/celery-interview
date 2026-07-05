"""Workbook conversion: .xlsx bytes -> per-sheet content.

Design decision (ADR 0001): we keep only cell content (one CSV per sheet),
not the original workbook — styles/formulas are irrelevant to sum and search.
"""

import csv
import io

from openpyxl import load_workbook

# Cells in search_text are joined with the ASCII "unit separator" control
# character — it cannot realistically appear in a search term or cell, so a
# term can never falsely match across two adjacent cells (a1,2b matching
# "1,2"), which a naive LIKE over the CSV content would allow.
CELL_SEPARATOR = "\x1f"


def workbook_to_sheets(data: bytes) -> list[tuple[str, str, float, str]]:
    """Parse .xlsx bytes into
    [(sheet_name, csv_content, numeric_sum, search_text), ...].

    Raises if the bytes are not a readable .xlsx — callers treat that as
    an invalid upload (ADR 0008: validity means openpyxl can open it).

    Design decision (ADR 0003): the numeric sum is derived here, while the
    workbook is open and openpyxl still exposes Excel's own cell typing —
    CSV erases types, so "what is a number" must be decided before storage.
    """
    # data_only=True: for formula cells, read the cached computed value —
    # we never evaluate formulas ourselves (out of scope per the PRD).
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        buf = io.StringIO()
        writer = csv.writer(buf)
        numeric_sum = 0.0
        cells = []
        for row in ws.iter_rows(values_only=True):
            writer.writerow("" if v is None else v for v in row)
            # A Number is what Excel types as numeric: text that merely looks
            # numeric (IDs, zip codes) is never summed. bool is excluded
            # explicitly because Python's bool subclasses int.
            numeric_sum += sum(
                v for v in row if isinstance(v, (int, float)) and not isinstance(v, bool)
            )
            # Search is case-insensitive (ADR 0004): lowercase once here so
            # queries never have to. Numeric cells participate via str().
            cells.extend(str(v).lower() for v in row if v is not None)
        search_text = CELL_SEPARATOR.join(cells)
        sheets.append((ws.title, buf.getvalue(), numeric_sum, search_text))
    return sheets
